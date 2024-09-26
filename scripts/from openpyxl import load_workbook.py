import glob
import os

from openpyxl import load_workbook


def rename_sheets(file_path):
    wb = load_workbook(file_path)
    rm = file_path.split("-")[1]

    for sheet in wb.sheetnames:
        new_name = f"RM_{rm}_{sheet}"
        wb[sheet].title = new_name

    wb.save(file_path)


# Define the data directory
data_dir = "/path/to/your/data/directory"

# Use this function in your main processing loop
for file in glob.glob(os.path.join(data_dir, "RM-*.xlsx")):
    rename_sheets(file)
